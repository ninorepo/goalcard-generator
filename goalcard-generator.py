from openpyxl import load_workbook
import shutil
from shutil import copyfile
from typing import Union
import sys

def extract_frontback(filepath, sheetname, target=None):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheetname]
    
    marker_col = None
    start_row = end_row = None
    panel_found = middle_found = False

    # Detect the column and rows for PANEL INSPECTION and MIDDLE INSPECTION
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if not panel_found and val.upper() == "PANEL INSPECTION":
                marker_col = cell.column
                start_row = cell.row
                panel_found = True
            elif panel_found and val.upper() == "MIDDLE INSPECTION" and cell.column == marker_col:
                end_row = cell.row
                middle_found = True
        if panel_found and middle_found:
            break

    if marker_col is None or start_row is None or end_row is None:
        print("Could not find PANEL INSPECTION and MIDDLE INSPECTION in same column.")
        return []

    # Find STD column (first column that contains 'STD' in its header)
    std_col = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and "STD" in str(cell.value).upper():
                std_col = cell.column
                break
        if std_col:
            break

    if std_col is None:
        print("STD column not found.")
        return []

    results = []

    # Iterate through the block including the markers
    for row in range(start_row, end_row + 1):
        op_cell = ws.cell(row=row, column=marker_col)
        std_cell = ws.cell(row=row, column=std_col)

        op_val = str(op_cell.value).strip() if op_cell.value else ""

        try:
            std_val = float(std_cell.value)
        except (TypeError, ValueError):
            continue  # skip if STD is not a number

        if not op_val:
            continue  # skip empty operation
        if std_val <= 0:
            continue  # skip zero or negative STD

        repeat = max(1, int(target // std_val)) if target else 1
        results.extend([(op_val, std_val)] * repeat)

    return results

def extract_assembly(filepath, sheetname, target=None):
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheetname]
    
    marker_col = None
    start_row = end_row = None
    middle_found = end_found = False

    # Step 1: Locate marker positions (must be in the same column)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if not middle_found and val.upper() == "MIDDLE INSPECTION":
                marker_col = cell.column
                start_row = cell.row
                middle_found = True
            elif middle_found and val.upper() == "END LINE INSPECTION" and cell.column == marker_col:
                end_row = cell.row
                end_found = True
        if middle_found and end_found:
            break

    if marker_col is None or start_row is None or end_row is None:
        print("Could not find MIDDLE INSPECTION and END LINE INSPECTION in the same column.")
        return []

    # Step 2: Find STD column (first cell with 'STD' in its value)
    std_col = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and "STD" in str(cell.value).upper():
                std_col = cell.column
                break
        if std_col:
            break

    if std_col is None:
        print("STD column not found.")
        return []

    results = []

    # Step 3: Loop from start_row + 1 (exclude MIDDLE INSPECTION) to end_row (include END LINE INSPECTION)
    for row in range(start_row + 1, end_row + 1):
        op_cell = ws.cell(row=row, column=marker_col)
        std_cell = ws.cell(row=row, column=std_col)

        op_val = str(op_cell.value).strip() if op_cell.value else ""

        try:
            std_val = float(std_cell.value)
        except (TypeError, ValueError):
            continue  # skip if STD is not a number

        if not op_val:
            continue  # skip empty operation
        if std_val <= 0:
            continue  # skip zero or negative STD

        repeat = max(1, int(target // std_val)) if target else 1
        results.extend([(op_val, std_val)] * repeat)

    return results
    


def inject_operations(operations, output_path, sheetname):
    if not operations:
        print(f"No operations to inject into '{sheetname}'. Skipping.")
        return

    wb = load_workbook(output_path)
    if sheetname not in wb.sheetnames:
        print(f"Sheet '{sheetname}' not found in '{output_path}'.")
        return

    ws = wb[sheetname]

    # Step 1: Locate gcstart and gcend in the same column
    marker_col = None
    start_row = end_row = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            val = str(cell.value).strip().lower() if cell.value else ""
            if val == "gcstart" and start_row is None:
                marker_col = cell.column
                start_row = cell.row
            elif val == "gcend" and marker_col and cell.column == marker_col and end_row is None:
                end_row = cell.row
        if start_row and end_row:
            break

    if not (start_row and end_row and marker_col):
        print("Start and End markers not found in the same column.")
        return

    total_rows = end_row - start_row + 1
    if len(operations) > total_rows:
        print(f"Warning: Only {total_rows} rows available, but {len(operations)} operations provided. Truncating.")
        operations = operations[:total_rows]

    # Step 2: Write operations starting from gcstart row
    current_row = start_row
    for op, std in operations:
        ws.cell(row=current_row, column=marker_col, value=op)
        ws.cell(row=current_row, column=marker_col + 1, value=std)
        current_row += 1

    # Step 3: Delete remaining empty rows from current_row to just before gcend
    deleted_rows = 0
    for row_idx in reversed(range(current_row, end_row)):
        val1 = ws.cell(row=row_idx, column=marker_col).value
        val2 = ws.cell(row=row_idx, column=marker_col + 1).value
        if (val1 is None or str(val1).strip() == "") and (val2 is None or str(val2).strip() == ""):
            ws.delete_rows(row_idx)
            deleted_rows += 1

    # Step 4: Recalculate gcend position and delete it
    updated_end_row = end_row - deleted_rows
    ws.delete_rows(updated_end_row)

    wb.save(output_path)
    print(f"Injected {len(operations)} operations starting from row {start_row}. Deleted leftover rows and removed gcend row.")



def template_copy(template_path, output_path):
    if template_path != output_path:
        copyfile(template_path, output_path)


def replace_marker_in_sheet(file_path: str, sheet_name: str, marker: str, new_value: Union[str, int, float]):
    """
    Replace the first occurrence of a marker keyword in a worksheet with a new value.
    
    Parameters:
    - file_path: Path to the Excel file (.xlsx)
    - sheet_name: Name of the sheet to search in
    - marker: The text in the cell to search for
    - new_value: The new value to replace it with (str, int, float)
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == marker:
                cell.value = new_value
                found = True
                break
        if found:
            break

    if not found:
        print(f"Marker '{marker}' not found in '{sheet_name}'.")
    else:
        wb.save(file_path)

def get_arg(index: int) -> str:
    """
    Returns the nth command-line argument (0-based, excluding script name).
    Raises IndexError if not enough arguments are provided.
    """
    try:
        return sys.argv[index + 1]  # +1 to skip the script name
    except IndexError:
        raise IndexError(f"Missing argument at index {index}.")

####################################################
engsheet = get_arg(0)
sheet = get_arg(1)

floor = get_arg(2)
cmt = get_arg(3)
tgt = float(get_arg(4))
date = get_arg(5)
output = get_arg(6)

print(engsheet)
print(sheet)
print(floor)
print(cmt)
print(tgt)
print(date)
print(output)

frontback = extract_frontback(engsheet, sheet, target=tgt)
assembly = extract_assembly(engsheet, sheet, target=tgt)

template_copy(template_path="gctemplate.xlsx", output_path=output)
# Inject into the template copy
inject_operations(
    operations=frontback,
    output_path=output,
    sheetname="frontback"
)
# Inject into the template copy
inject_operations(
    operations=assembly,
    output_path=output,
    sheetname="assembly"
)

replace_marker_in_sheet(output, "frontback", "gcfloor", floor)
replace_marker_in_sheet(output, "frontback", "gccmt", cmt)
replace_marker_in_sheet(output, "frontback", "gcdate", date)
replace_marker_in_sheet(output, "assembly", "gcfloor", floor)
replace_marker_in_sheet(output, "assembly", "gccmt", cmt)
replace_marker_in_sheet(output, "assembly", "gcdate", date)

#python goalcard-generator.py engsheet.xlsx 592 "LINE 4/B8" "UQLK 312" 80 "17 Juli 2025" "gcoutput.xlsx"